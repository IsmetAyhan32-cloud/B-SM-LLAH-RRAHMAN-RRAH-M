import json
import os
import uuid
from typing import List, Dict, Any

from flask import Flask, render_template, request, redirect, url_for, flash, session as flask_session
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = os.environ.get("ATTENDANCE_APP_SECRET", "dev-secret-key")

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
SESSIONS_FILE = os.path.join(DATA_DIR, "sessions.json")
TMP_UPLOAD_DIR = os.environ.get("TMPDIR", "/tmp")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(TMP_UPLOAD_DIR, exist_ok=True)


def load_sessions() -> List[Dict[str, Any]]:
    if not os.path.exists(SESSIONS_FILE):
        return []
    with open(SESSIONS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_sessions(sessions: List[Dict[str, Any]]):
    with open(SESSIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(sessions, f, ensure_ascii=False, indent=2)


def parse_excel(file_path: str) -> List[Dict[str, Any]]:
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    students = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        number = str(row[0]).strip() if row[0] is not None else ""
        name = str(row[1]).strip() if row[1] is not None else ""
        department = str(row[2]).strip() if row[2] is not None else ""
        weeks_raw = row[3:17]
        attendance: List[bool] = []
        for cell in weeks_raw:
            value = str(cell).strip() if cell is not None else ""
            attendance.append(value == "+")
        if len(attendance) > 14:
            attendance = attendance[:14]
        while len(attendance) < 14:
            attendance.append(False)
        students.append(
            {
                "number": number,
                "name": name,
                "department": department,
                "attendance": attendance,
            }
        )
    return students


@app.route("/")
def index():
    return redirect(url_for("teacher_panel"))


@app.route("/teacher", methods=["GET"])
def teacher_panel():
    sessions = load_sessions()
    return render_template("teacher.html", sessions=sessions, weeks=list(range(1, 15)))


@app.route("/teacher/create", methods=["POST"])
def create_session():
    session_name = request.form.get("session_name", "").strip()
    excel_file = request.files.get("excel_file")

    if not session_name:
        flash("Oturum adı gereklidir.", "error")
        return redirect(url_for("teacher_panel"))

    if not excel_file:
        flash("Excel dosyası yüklenmelidir.", "error")
        return redirect(url_for("teacher_panel"))

    filename = secure_filename(excel_file.filename)
    if not filename:
        flash("Geçerli bir dosya adı gereklidir.", "error")
        return redirect(url_for("teacher_panel"))

    temp_filename = f"{uuid.uuid4()}-{filename}"
    temp_path = os.path.join(TMP_UPLOAD_DIR, temp_filename)
    excel_file.save(temp_path)

    try:
        students = parse_excel(temp_path)
    except Exception as exc:  # pylint: disable=broad-except
        flash(f"Excel dosyası okunamadı: {exc}", "error")
        os.remove(temp_path)
        return redirect(url_for("teacher_panel"))

    os.remove(temp_path)

    sessions = load_sessions()
    new_session = {
        "id": str(uuid.uuid4()),
        "name": session_name,
        "active_week": None,
        "students": students,
    }
    sessions.append(new_session)
    save_sessions(sessions)

    flash("Oturum başarıyla oluşturuldu.", "success")
    return redirect(url_for("teacher_panel"))


@app.route("/teacher/<session_id>/week", methods=["POST"])
def update_active_week(session_id: str):
    week_value = request.form.get("active_week")
    sessions = load_sessions()

    for session_data in sessions:
        if session_data["id"] == session_id:
            if week_value:
                try:
                    week_number = int(week_value)
                    if week_number < 1 or week_number > 14:
                        raise ValueError
                    session_data["active_week"] = week_number
                except ValueError:
                    flash("Hafta değeri 1-14 arasında olmalıdır.", "error")
                    save_sessions(sessions)
                    return redirect(url_for("teacher_panel"))
            else:
                session_data["active_week"] = None
            break

    save_sessions(sessions)
    flash("Aktif hafta güncellendi.", "success")
    return redirect(url_for("teacher_panel"))


@app.route("/student", methods=["GET", "POST"])
def student_login():
    if request.method == "POST":
        student_number = request.form.get("student_number", "").strip()
        if not student_number:
            flash("Öğrenci numarası gereklidir.", "error")
            return redirect(url_for("student_login"))

        flask_session["student_number"] = student_number
        flash("Giriş başarılı.", "success")
        return redirect(url_for("student_sessions"))

    return render_template("student_login.html")


@app.route("/student/sessions", methods=["GET"])
def student_sessions():
    student_number = flask_session.get("student_number")
    if not student_number:
        flash("Lütfen önce giriş yapın.", "error")
        return redirect(url_for("student_login"))

    sessions = [session for session in load_sessions() if session.get("active_week")]
    return render_template("student_sessions.html", sessions=sessions, student_number=student_number)


@app.route("/student/sessions/<session_id>/attend", methods=["POST"])
def attend_session(session_id: str):
    student_number = flask_session.get("student_number")
    if not student_number:
        flash("Giriş yapılmadı.", "error")
        return redirect(url_for("student_login"))

    sessions = load_sessions()
    target_session = next((session for session in sessions if session["id"] == session_id), None)

    if not target_session:
        flash("Oturum bulunamadı.", "error")
        return redirect(url_for("student_sessions"))

    active_week = target_session.get("active_week")
    if not active_week:
        flash("Bu oturum için aktif hafta ayarlanmamış.", "error")
        return redirect(url_for("student_sessions"))

    student_entry = next((student for student in target_session["students"] if student["number"] == student_number), None)
    if not student_entry:
        flash("Öğrenci listede bulunamadı.", "error")
        return redirect(url_for("student_sessions"))

    week_index = active_week - 1
    student_entry["attendance"][week_index] = True
    save_sessions(sessions)

    flash("Yoklama işleminiz alındı.", "success")
    return redirect(url_for("student_sessions"))


@app.context_processor
def inject_enumerate():
    return {"enumerate": enumerate}


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
